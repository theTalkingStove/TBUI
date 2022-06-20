#GUI

# #Import tkinter library
# from tkinter import *
# #Create an instance of tkinter frame or window
# win= Tk()
# #Set the geometry of tkinter frame
# win.geometry("720x720")
# #Create a Label widget
# label= Label(win, text="Open Source Learning is Awesome!", font= ('Courier 20 underline'))
# label.pack()

# win.mainloop()
# -----------------------------------------------------------------

from tkinter import *
from openpyxl import *
from openpyxl import Workbook
from openpyxl.cell import cell
from openpyxl import load_workbook



win=Tk()


#creating multiple Entrys
entries = []
address=[]

for i in range(5):
    Entry(win).grid(row=0, column=i+1)
    Entry(win).focus
    entries.append(Entry(win))

# #making the command function for the "Initialize" button so we can ask for the addresses

def InstrCheck():
    for address in entries:
        Entry(win) == address
    else:
        print('address incorrect')

#button with it's features and position changed
bt = Button(text="Initialize", width=25, height=5, bg="red", fg="white", padx=55, pady=33, command=InstrCheck)
bt.grid(row=10, column=10)

#making a button for the data to be stored into the excel file
#what data do we need?
#what are the formulas used/needed?

def ExcelDriver():


    wb = load_workbook(r'C:\Users\EngLab\Documents\ALL_MISC_TESTS\testing_input.xlsx')
    ws = wb.active
    a = entries[0].get
    b = entries[1].get
    c = entries[2].get
    d = entries[3].get

    ws.cell(row = 1, column = 1, value =a)
    ws.cell(row = 1, column = 2, value =b)
    ws.cell(row = 1, column = 3, value =c)
    ws.cell(row = 1, column = 4, value =d)
        
       




    wb.save(r'C:\Users\EngLab\Documents\ALL_MISC_TESTS\testing_input.xlsx')



bt2 = Button(text="OpenExcelInput", width=25, height=5, bg="blue", fg="green", padx=35, pady=30, command=ExcelDriver())
bt2.grid(row=10, column=11)


























win.mainloop()

